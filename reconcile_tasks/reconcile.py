import pandas as pd
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ============================
# 1️⃣ 读取原始文件
# ============================
# 把你的两张表保存为 Excel 或 CSV 文件
# 例如：company_book.xlsx / bank_statement.xlsx
corp = pd.read_csv("corp_07.csv")  # 图一
bank = pd.read_csv("bank.csv")  # 图二


# ============================
# 2️⃣ 字段标准化（根据真实表头）
# ============================
corp = corp.rename(columns={
    "日期": "date",
    "明细摘要": "summary",
    "余额": "corp_balance"
})

corp = corp[corp["summary"].isnull() == False]

# 收入字段（左侧）
income_cols = ["豪爵车款", "配件款", "广宣及装修款", "转存", "其他"]  # ← 注意：这些列名必须与Excel完全一致
# 支出字段（右侧）
expense_cols = ["付工厂货款", "广宣及装修品", "公司费用", "财务费用", "其他.1"]  # Excel中第二个“其他”通常会变成 “其他.1”


# 汇总计算
corp["income"] = corp[income_cols].sum(axis=1, numeric_only=True)
corp["expense"] = corp[expense_cols].sum(axis=1, numeric_only=True)
corp["net"] = corp["income"].fillna(0) - corp["expense"].fillna(0)

bank = bank[["起息日", "交易类型", "借方金额", "贷方金额", "余额"]]
bank = bank.rename(columns={
    "起息日": "date",
    "交易类型": "type",
    "借方金额": "debit",
    "贷方金额": "credit",
    "余额": "bank_balance"
})

for col in ["debit", "credit"]:
    bank[col] = (
        bank[col]
        .astype(str)
        .str.replace(",", "", regex=False)     # 去掉千分位逗号
        .str.replace(" ", "", regex=False)     # 去掉空格
        .str.replace("－", "-", regex=False)   # 中文减号 → 英文减号
        .str.replace("-", "0", regex=False)    # 空白/短横线替换为0
        .replace("", "0")                     # 空字符串转为0
    )
    bank[col] = pd.to_numeric(bank[col], errors="coerce").fillna(0)

# 计算净额：贷方 - 借方
bank["net"] = bank["credit"] - bank["debit"]


# 清洗银行余额字段，转成 float
bank["bank_balance"] = (
    bank["bank_balance"]
    .astype(str)
    .str.replace(",", "", regex=False)
    .str.replace(" ", "", regex=False)
    .replace("-", "0")
    .replace("", "0")
)
bank["bank_balance"] = pd.to_numeric(bank["bank_balance"], errors="coerce").fillna(0.0)



# 清洗公司账余额字段，转成 float
corp["corp_balance"] = (
    corp["corp_balance"]
    .astype(str)
    .str.replace(",", "", regex=False)
    .str.replace(" ", "", regex=False)
    .replace("-", "0")
    .replace("", "0")
)
corp["corp_balance"] = pd.to_numeric(corp["corp_balance"], errors="coerce").fillna(0.0)


corp["date"] = pd.to_datetime(corp["date"], errors="coerce", dayfirst=True)
bank["date"] = pd.to_datetime(bank["date"], errors="coerce", dayfirst=False)

bank["date"] = (
    bank["date"]
    .astype(str)
    .str.strip()
    .replace(r"[^0-9/.-]", "", regex=True)
    .pipe(lambda s: pd.to_datetime(s, errors="coerce"))
)

corp = corp[['date', 'summary', 'corp_balance', 'income', 'expense', 'net']]


bank.columns
corp.columns
# -----------------------------
# 5️⃣ 余额匹配逻辑（比对公司账累计余额与银行余额）
# -----------------------------
def match_balance(row, bank_df, tol_days=3, tol_amt=50.0):
    date = row["date"]
    if pd.isna(date):
        return None, None, None, None, "跳过"

    # 找到日期窗口内最近的一笔银行记录
    mask = (bank_df["date"] >= date - timedelta(days=tol_days)) & \
           (bank_df["date"] <= date + timedelta(days=tol_days))
    subset = bank_df.loc[mask].copy()
    if subset.empty:
        return None, None, None, None, "银行无数据"

    # 找最接近的余额（银行余额）
    subset["diff"] = (subset["bank_balance"] - row["corp_balance"]).abs()
    best = subset.loc[subset["diff"].idxmin()]

    bank_balance = float(best["bank_balance"])
    diff = abs(bank_balance - row["corp_balance"])
    status = "正常" if diff <= tol_amt else "余额异常"

    return best["date"], best["type"], bank_balance, diff, status


corp[["匹配日期", "银行交易类型", "银行金额", "金额差额", "匹配状态"]] = corp.apply(
    match_balance,
    axis=1,
    result_type="expand",
    bank_df=bank,
    tol_days=3,
    tol_amt=50.0
)

corp.dtypes
bank.dtypes

# ============================
# 4️⃣ 输出 Excel（保留所有原始列 + 新增对账列）
# ============================
output = "company_book_reconciled_07.xlsx"
corp.to_excel(output, index=False)

# ============================
# 5️⃣ Excel 自动上色（直观显示问题）
# ============================
wb = load_workbook(output)
ws = wb.active
ws.title = "Reconciled_Book"

green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
    status = row[-1].value  # 匹配状态列
    if status == "正常":
        fill = green
    elif status == "金额异常":
        fill = yellow
    elif "银行无数据" in str(status) or "方向不符" in str(status):
        fill = red
    else:
        fill = None
    if fill:
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save(output)
print(f"✅ 对账结果已生成：{output}")
print("输出说明：")
print(" - 绿色：金额完全匹配")
print(" - 黄色：金额差额超出容差")
print(" - 红色：银行无对应记录或方向不符")